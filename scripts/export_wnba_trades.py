#!/usr/bin/env python
"""Export the operator's WNBA fills to a plain spreadsheet for hand annotation.

    python scripts/export_wnba_trades.py

Writes ``<MERIDIAN_DATA_DIR>/exports/wnba-trades-<date>.xlsx`` and a ``.csv``
twin with the same rows. One row per fill, flat columns, three empty columns
at the end for the operator to fill in by hand.

**Read-only.** The only venue calls are :class:`PolymarketAuthedClient` GETs
(the class can express no other verb) and the free public settlement endpoint.
Nothing here places, modifies or cancels an order.

**It will not overwrite an existing sheet.** Once the operator has written in
the annotation columns, that file is the only copy of those notes and is not
regenerable — so a second run on the same day stops rather than clobbering it.
Pass ``--force`` to overwrite deliberately, or ``--out`` to write elsewhere.

Reading a row
-------------
``market`` is the contract traded, as the venue labels it; ``position`` is the
exposure that fill put on. They differ whenever a contract was SOLD — selling
the Under is being long the Over — so read ``position`` for what was actually
backed. Money columns describe only the exposure the row itself opened: a pure
exit leaves them blank, because its proceeds are booked on the entry it closed,
which is also the row whose annotation explains the trade.

The parsing and P&L arithmetic live in :mod:`core.audit.wnba_trade_sheet`,
which documents the two venue conventions this depends on and is covered by
``tests/test_wnba_trade_sheet.py``. Read that module's docstring before
trusting a number in the sheet.

Fees and rebates are NOT netted into P&L
----------------------------------------
The ``fees`` column is the venue's per-execution commission as reported —
positive when we paid it as the aggressor, negative when we were the passive
side and earned a rebate. Account-level ``ACTIVITY_TYPE_TAKER_FEE_REBATE``
credits carry no trade or order id (only a date and an opaque uuid), so they
cannot honestly be attributed to a fill; their total is printed in the run
summary and left out of the sheet rather than spread across rows on a guess.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import sys
import time
from decimal import Decimal
from pathlib import Path

# Running `python scripts/export_wnba_trades.py` puts scripts/ on sys.path, not
# the repo root, so `core` would resolve to whichever checkout the venv has
# installed — the wrong one inside a git worktree. Anchor it to THIS file.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import structlog

from core import paths
from core.audit.wnba_trade_sheet import (
    CENTRAL,
    UTC,
    Resolution,
    SheetRow,
    WnbaFill,
    build_rows,
    parse_activity,
)

log = structlog.get_logger(__name__)

ACTIVITIES_PATH = "/v1/portfolio/activities"
PAGE_LIMIT = 100
#: The whole account history is a few hundred events (681 on 2026-08-17, seven
#: pages). The cap is generous so a longer history still walks to the end, and
#: the pause keeps us well under the authenticated host's ~5 req/s throttle.
MAX_PAGES = 200
PAGE_PAUSE_SECONDS = 0.35

#: (header, width). The last three are deliberately empty — they are the
#: reason the sheet exists.
COLUMNS: tuple[tuple[str, int], ...] = (
    ("date/time (UTC)", 20),
    ("date/time (CT)", 20),
    ("game", 24),
    ("market", 16),
    ("position", 16),
    ("side", 6),
    ("price", 8),
    ("size", 9),
    ("role", 11),
    ("$ in", 10),
    ("$ out/settled", 14),
    ("P&L (FIFO)", 12),
    ("closed by", 11),
    ("fees", 8),
    ("placed by", 11),
    ("market slug", 38),
    ("why I made this trade", 42),
    ("what I'd do differently", 42),
    ("pattern name", 24),
)


def _money(value: Decimal | None) -> float | None:
    return None if value is None else float(round(value, 2))


def row_values(row: SheetRow, placed_by: str) -> list:
    """One sheet row. Money is rounded here, once, so the .xlsx and the .csv
    cannot drift apart.

    A row that opened no exposure of its own (a pure exit) leaves the money
    columns BLANK — its proceeds are booked on the entry it closed. A row that
    did open exposure always prints a number, including ``0.00``: a position
    that settled worthless returned zero dollars, which is a result, and a
    blank cell would read as "nothing happened" instead.
    """
    fill = row.fill
    has_lot = row.opened > 0
    return [
        fill.at.astimezone(UTC).strftime("%Y-%m-%d %H:%M:%S"),
        fill.at.astimezone(CENTRAL).strftime("%Y-%m-%d %H:%M:%S"),
        fill.game,
        fill.market,
        row.position,
        "BUY" if fill.is_buy else "SELL",
        float(round(fill.yes_price, 4)),
        float(fill.shares),
        row.role,
        _money(row.dollars_in) if has_lot else None,
        _money(row.dollars_out) if has_lot else None,
        _money(row.pnl),
        row.closed_by,
        _money(fill.commission),
        placed_by,
        fill.market_slug,
        "", "", "",
    ]


# --------------------------------------------------------------------------- #
# Data acquisition — GET only
# --------------------------------------------------------------------------- #


def activities_from_export(path: Path) -> tuple[list[dict], str]:
    """Read a pinned venue export instead of calling the venue.

    Accepts both shapes this project produces:

    * a flat ``[activity, ...]`` list, as :func:`fetch_activities` returns;
    * the paged envelope ``{"pages": [{"activities": [...], "eof": bool}],
      "fetched_at": "..."}`` that the venue-export tooling writes to
      ``backups/exports/``.

    **Completeness is checked, not assumed.** The live path guarantees it by
    walking to ``eof``; reading a file bypasses that walk entirely, so a
    truncated snapshot would produce a sheet silently missing the OLDEST trades
    — the same defect ``fetch_activities`` raises on, arriving through a door
    that had no guard. A snapshot whose last page never reached eof is refused.

    Returns the activities and a provenance string naming the snapshot, so the
    sheet can say which pinned artifact it was generated from.
    """
    raw = json.loads(path.read_text())

    if isinstance(raw, list):
        # A flat dump carries no eof marker, so completeness cannot be checked
        # here. Say so rather than implying it was verified.
        return ([a for a in raw if isinstance(a, dict)],
                f"{path.name} (flat dump — completeness not verifiable)")

    if not isinstance(raw, dict) or not isinstance(raw.get("pages"), list):
        raise RuntimeError(
            f"{path}: not a venue activities export — expected a list of "
            "activities or an object with a 'pages' array"
        )

    pages = raw["pages"]
    if not pages:
        raise RuntimeError(f"{path}: export contains no pages")

    last = pages[-1] if isinstance(pages[-1], dict) else {}
    if not (last.get("eof") or not last.get("nextCursor")):
        raise RuntimeError(
            f"{path}: the export never reached eof — its last page still has a "
            "nextCursor, so the OLDEST trades are missing and any sheet built "
            "from it would be quietly incomplete. Re-take the export."
        )

    activities = [a for page in pages if isinstance(page, dict)
                  for a in (page.get("activities") or []) if isinstance(a, dict)]
    at = raw.get("fetched_at") or "unknown time"
    return activities, f"{path.name} ({len(activities)} activities, fetched {at})"


def fetch_activities(client) -> list[dict]:
    """Walk the activity feed to eof. Completeness over speed."""
    out: list[dict] = []
    cursor: str | None = None
    for page in range(MAX_PAGES):
        params: dict = {"limit": PAGE_LIMIT}
        if cursor:
            params["cursor"] = cursor
        resp = client.get(ACTIVITIES_PATH, params=params)
        if resp.status_code != 200:
            raise RuntimeError(
                f"activities HTTP {resp.status_code}: {resp.body_text[:200]}"
            )
        body = json.loads(resp.body_text)
        out.extend(a for a in (body.get("activities") or []) if isinstance(a, dict))
        cursor = body.get("nextCursor")
        if body.get("eof") or not cursor:
            log.info("activities_complete", pages=page + 1, activities=len(out))
            return out
        time.sleep(PAGE_PAUSE_SECONDS)
    raise RuntimeError(
        f"activity feed did not reach eof in {MAX_PAGES} pages — the sheet would "
        "be missing the oldest trades. Raise MAX_PAGES and re-run."
    )


def settlement_lookup(gateway):
    """``slug -> Decimal(0|1) | None`` from the public settlement endpoint.

    Ground truth over inference: the resolution activity's before/after
    bookkeeping has undocumented sign conventions, so it is not used to derive
    a payout. Unknown stays None and the lot is reported unscored.
    """
    cache: dict[str, Decimal | None] = {}

    def lookup(slug: str) -> Decimal | None:
        if slug in cache:
            return cache[slug]
        try:
            value = gateway.get_settlement(slug).get("settlement")
            result = Decimal(value) if value in (0, 1) else None
        except Exception as exc:                                  # noqa: BLE001
            log.warning("settlement_fetch_failed", market=slug, error=str(exc)[:120])
            result = None
        cache[slug] = result
        return result

    return lookup


def button_order_ids() -> tuple[set[str] | None, dt.datetime | None]:
    """System-placed venue order ids, and **how current the source is**.

    Attribution rests on a completeness-by-construction argument, not on the
    set being non-empty: ``core/api.py`` writes the order row and commits it
    **before** calling the venue (``s.add`` / ``s.commit`` at api.py:1606, submit
    at :1649), so every order this system sent has a row. Absence of a matching
    id is therefore evidence the order was not ours, and ``hand`` is a sound
    label rather than a lucky one.

    **The argument covers accepted orders, not every row.** ``venue_order_id``
    is written back only after a *successful* submit (api.py:1670), so an
    ambiguously-failed submission leaves a row with **no id** and is outside
    this set. If such an order did reach the venue and fill, its fill would
    carry an id we never recorded and would be labelled ``hand`` — the one case
    this argument does not cover. Measured 2026-08-26: **5 rows, 5 with ids, 0
    without**, so the gap is real in principle and empty in practice. It is
    written here rather than left in a review comment because this docstring is
    what someone will lean on in a year.

    That argument holds **only if this database is the one those writes went
    to.** Read against a lagging copy the rows are simply missing, and every
    system order becomes a hand trade on the operator's own annotation sheet.

    So currency is asserted, not assumed — and the witness is deliberately NOT
    ``max(created_at)`` of ``orders``. That conflates *when the table last had
    something to say* with *how current the source is*: this system has placed 5
    orders ever, all on 2026-08-07, so a complete and perfectly current table
    looks nineteen days stale by that measure. The witness is instead the
    highest-frequency writer in the system, ``market_snapshots``, whose horizon
    bounds how recent ANY of this database's contents are.

    Returns ``(None, None)`` when the source cannot be read or dated at all —
    the caller must then attribute nothing.
    """
    try:
        from sqlalchemy import text as sql

        from core.storage import get_engine

        with get_engine().connect() as conn:
            rows = conn.execute(
                sql("select venue_order_id from orders where venue_order_id is not null")
            ).all()
            current_through = conn.execute(
                sql("select max(captured_at) from market_snapshots")
            ).scalar()
        if current_through is None:
            log.warning("orders_source_undateable")
            return None, None
        return {str(r[0]) for r in rows}, current_through
    except Exception as exc:                                      # noqa: BLE001
        log.warning("button_order_ids_unavailable", error=str(exc)[:120])
        return None, None


def export_fetched_at(path: Path) -> dt.datetime | None:
    """``fetched_at`` from a pinned export envelope, or None for a flat dump."""
    raw = json.loads(path.read_text())
    stamp = raw.get("fetched_at") if isinstance(raw, dict) else None
    if not stamp:
        return None
    try:                              # the venue tooling writes 20260825T233057Z
        return dt.datetime.strptime(str(stamp), "%Y%m%dT%H%M%SZ").replace(tzinfo=dt.timezone.utc)
    except ValueError:
        return _parse_ts(stamp)


def market_live_through() -> dt.datetime | None:
    """When live market activity was last OBSERVED — an anchor external to the
    export.

    ``max(captured_at) where is_live`` is the recorder's last live tick, which
    approximates the newest game's final whistle. It comes from the recorder,
    not from the activities feed, so comparing it against an export's
    ``fetched_at`` has **two real operands**.

    That property is the whole point. The first version of this preflight
    compared the export's timestamp against the latest fill *in that export* —
    and every fill necessarily predates the fetch, so the check could not fail.
    It passed its own mutation test only because the test injected the two
    values independently while the live call path derived one from the other.
    A short window cannot be detected from inside the window.

    One-sided by nature: if this source is itself behind, it under-reports and
    cannot prove shortness. It can only ever say "the market was live after your
    snapshot, so your snapshot is short" — never "your snapshot is complete".
    """
    try:
        from sqlalchemy import text as sql

        from core.storage import get_engine

        with get_engine().connect() as conn:
            return conn.execute(
                sql("select max(captured_at) from market_snapshots where is_live")
            ).scalar()
    except Exception as exc:                                      # noqa: BLE001
        log.warning("market_live_through_unavailable", error=str(exc)[:120])
        return None


def preflight(latest_fill: dt.datetime,
              export_at: dt.datetime | None,
              covers_through: dt.datetime | None,
              attribution_through: dt.datetime | None,
              gateway_url: str) -> tuple[bool, list[str]]:
    """Refuse to build a sheet whose sources cannot cover the trades in it.

    The sheet's window is the **oldest** of its sources, not the newest. A leg
    that predates the trading produces a sheet that is silently short while
    passing every internal check: the export walks to ``eof`` correctly over a
    window that ended too early, and the arithmetic is sound on whatever it was
    handed. ``eof`` proves the READ; only ``covers_through`` proves the WINDOW.

    ``covers_through`` must come from OUTSIDE the export — the recorder's last
    live tick, or a caller's explicit ``--covers-through`` claim. Without it the
    window is unbounded, and an unbounded window is exactly what looks finished.

    Returns ``(ok, lines)``. ``ok=False`` means stop, not warn.
    """
    lines: list[str] = []
    ok = True

    lines.append(f"  latest fill in this sheet   {latest_fill}   (informational — "
                 "derived FROM the export, so it cannot bound it)")

    if export_at is None:
        lines.append("  activities export           UNDATED (flat dump) — cannot bound")
        ok = False
    else:
        lines.append(f"  activities export           {export_at}")

    if covers_through is None:
        lines.append("  market live through         UNKNOWN — no external anchor, so "
                     "the window\n                              cannot be bounded. "
                     "Pass --covers-through to assert it.")
        ok = False
    else:
        short = export_at is not None and covers_through > export_at
        lines.append(f"  market live through         {covers_through}"
                     f"{'   *** LIVE AFTER THE EXPORT — it is short' if short else ''}")
        ok &= not short

    live = gateway_url.rstrip("/") == "https://gateway.polymarket.us"
    lines.append(f"  settlement gateway          {gateway_url}"
                 f"{'' if live else '   *** NOT THE LIVE GATEWAY'}")
    ok &= live

    if attribution_through is None:
        lines.append("  attribution database        UNREADABLE — rows will read unknown")
    else:
        lag = attribution_through < latest_fill
        lines.append(f"  attribution database        {attribution_through}"
                     f"{'   *** rows after this read unknown' if lag else ''}")
        # Not fatal: `unknown` degrades a column, a short export drops rows.

    return ok, lines


def placed_by(fill_at: dt.datetime, order_id: str,
              ours: set[str] | None, source_current_through: dt.datetime | None) -> str:
    """``system`` | ``hand`` | ``unknown``.

    ``unknown`` only where the source genuinely cannot speak: it was unreadable,
    or the fill is **newer than anything in the database**, so its silence about
    that order carries no information.
    """
    if ours is None or source_current_through is None:
        return "unknown"
    if order_id in ours:
        return "system"
    if fill_at > source_current_through:
        return "unknown"
    return "hand"


# --------------------------------------------------------------------------- #
# Writing
# --------------------------------------------------------------------------- #


def write_csv(path: Path, rows: list[list]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow([header for header, _ in COLUMNS])
        writer.writerows(rows)


def write_xlsx(path: Path, rows: list[list]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    book = Workbook()
    sheet = book.active
    sheet.title = "WNBA trades"
    sheet.append([header for header, _ in COLUMNS])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for index, (_, width) in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in rows:
        sheet.append(row)
    book.save(path)


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #


def _collect(activities: list[dict]) -> tuple[list[WnbaFill], list[Resolution], int]:
    fills: list[WnbaFill] = []
    resolutions: list[Resolution] = []
    unparsed = 0
    for raw in activities:
        fill, resolution, ok = parse_activity(raw)
        if not ok:
            unparsed += 1
            log.warning("unparsed_activity", type=raw.get("type"),
                        keys=sorted(raw.keys())[:10])
        if fill is not None:
            fills.append(fill)
        if resolution is not None:
            resolutions.append(resolution)
    fills.sort(key=lambda f: (f.at, f.market_slug))
    return fills, resolutions, unparsed


def _rebate_total(activities: list[dict]) -> Decimal:
    total = Decimal("0")
    for raw in activities:
        if raw.get("type") != "ACTIVITY_TYPE_TAKER_FEE_REBATE":
            continue
        value = ((raw.get("accountBalanceChange") or {}).get("amount") or {}).get("value")
        if value is not None:
            total += Decimal(str(value))
    return total


def main() -> int:
    parser = argparse.ArgumentParser(prog="meridian-export-wnba-trades")
    parser.add_argument("--out", type=Path, default=None,
                        help="explicit .xlsx path (default: "
                             "$MERIDIAN_DATA_DIR/exports/wnba-trades-<today>.xlsx)")
    parser.add_argument("--force", action="store_true",
                        help="overwrite an existing sheet — DESTROYS any hand "
                             "annotations already in it")
    parser.add_argument("--activities-json", type=Path, default=None,
                        help="build from a PINNED venue export instead of calling "
                             "the venue. Takes either a flat activities list or "
                             "the paged envelope under backups/exports/. Makes the "
                             "sheet reproducible: the same file always yields the "
                             "same sheet, so it stays re-gradable against the "
                             "ledger it was checked against")
    parser.add_argument("--covers-through", type=dt.datetime.fromisoformat, default=None,
                        help="ISO instant the caller ASSERTS this export covers "
                             "through, when no recorder data is available to anchor "
                             "it. A claim someone makes, rather than one the export "
                             "makes about itself.")
    args = parser.parse_args()
    if args.covers_through and args.covers_through.tzinfo is None:
        args.covers_through = args.covers_through.replace(tzinfo=dt.timezone.utc)

    import logging
    logging.basicConfig(format="%(message)s", stream=sys.stderr, level=logging.INFO)

    import core.storage  # noqa: F401  loads .env, as every CLI here does

    provenance = None
    if args.activities_json:
        activities, provenance = activities_from_export(args.activities_json)
    else:
        from core.polymarket.client import PolymarketAuthedClient, USCredentials

        client = PolymarketAuthedClient(USCredentials.from_env())
        try:
            activities = fetch_activities(client)
        finally:
            client.close()

    fills, resolutions, unparsed = _collect(activities)
    if not fills:
        print("No WNBA fills on this account — nothing to write.", file=sys.stderr)
        return 1

    from core.polymarket.client import PolymarketGatewayClient

    with PolymarketGatewayClient() as gateway:
        rows, unscored = build_rows(fills, resolutions, settlement_lookup(gateway))

    ours, current_through = button_order_ids()

    # Preflight: the sheet's window is the OLDEST of its sources. Checked before
    # anything is written, and fatal rather than advisory for a short export —
    # a sheet missing a game looks finished.
    from core.polymarket.client import RECORDER

    ok, lines = preflight(
        latest_fill=max(f.at for f in fills),
        export_at=export_fetched_at(args.activities_json) if args.activities_json else None,
        covers_through=args.covers_through or market_live_through(),
        attribution_through=current_through,
        gateway_url=RECORDER.gateway_base_url,
    )
    print("\n  source coverage:", file=sys.stderr)
    for line in lines:
        print(line, file=sys.stderr)
    if not ok:
        print("\n  REFUSING to write: a source does not cover the trades in this "
              "sheet.\n  Re-take the export (or fix the gateway) and re-run. A short "
              "sheet\n  looks finished, which is worse than no sheet.", file=sys.stderr)
        return 1
    values = [
        row_values(row,
                   placed_by(row.fill.at, row.fill.venue_order_id, ours, current_through))
        for row in rows
    ]
    unknown = sum(1 for v in values if "unknown" in v)
    if unknown:
        print(f"\n  !! {unknown} of {len(values)} rows have placed-by UNKNOWN. The "
              f"attribution\n     database is current only through {current_through}, "
              "so its silence about\n     newer orders carries no information. Point "
              "at a current database to\n     resolve them — they are NOT hand trades "
              "by default.", file=sys.stderr)

    # The operator's local date, not UTC. Run at 19:10 CT on 17 Aug, `now(UTC)`
    # is already the 18th, and a sheet of tonight's trading would be filed under
    # tomorrow. This is an operator-facing artifact; it follows the operator's
    # clock. Row timestamps still carry both zones.
    today = dt.datetime.now(CENTRAL).date().isoformat()
    xlsx = args.out or (paths.data_dir() / "exports" / f"wnba-trades-{today}.xlsx")
    csv_path = xlsx.with_suffix(".csv")
    for path in (xlsx, csv_path):
        if path.exists() and not args.force:
            print(f"REFUSING to overwrite {path}\n"
                  "  It may already hold hand-written annotations, which are not "
                  "regenerable.\n  Move it aside, pass --out, or --force to "
                  "overwrite deliberately.", file=sys.stderr)
            return 2
    xlsx.parent.mkdir(parents=True, exist_ok=True)
    write_xlsx(xlsx, values)
    write_csv(csv_path, values)

    # ---- summary -------------------------------------------------------- #
    first, last = rows[0].fill.at, rows[-1].fill.at
    staked = sum((r.dollars_in for r in rows), Decimal("0"))
    returned = sum((r.dollars_out for r in rows), Decimal("0"))
    scored = [r for r in rows if r.pnl is not None]
    realized = sum((r.pnl for r in scored), Decimal("0"))
    fees = sum((r.fill.commission for r in rows), Decimal("0"))
    rebates = _rebate_total(activities)

    print(f"\nWNBA trade sheet — {len(values)} rows (one per fill)")
    print("=" * 68)
    print(f"  date range (UTC)  {first:%Y-%m-%d %H:%M} -> {last:%Y-%m-%d %H:%M}")
    print(f"  date range (CT)   {first.astimezone(CENTRAL):%Y-%m-%d %H:%M} -> "
          f"{last.astimezone(CENTRAL):%Y-%m-%d %H:%M}")
    print(f"  activities walked {len(activities)}  ({unparsed} unparsed)")
    print(f"  markets           {len({r.fill.market_slug for r in rows})}")
    print(f"  staked            ${staked:,.2f}")
    print(f"  returned          ${returned:,.2f}")
    print(f"  realized P&L      ${realized:,.2f}   ({len(scored)} of {len(rows)} "
          f"rows fully closed and scored)")
    print(f"  fees as reported  ${fees:,.2f}   (per-fill, NOT netted into P&L)")
    print(f"  taker fee rebates ${rebates:,.2f}   (account-level, unattributable "
          "to a fill; not in the sheet.\n                     50% of that ET-day's own "
          "taker fees -- a promo that ENDED 2026-05-10, V24)")
    # The caveat prints next to the number, not in a docstring nobody opens:
    # an unqualified total is what gets quoted, whatever the prose says.
    print("\n  !! P&L HERE IS FIFO, PER ROUND TRIP. The venue's own realizedPnl is "
          "a\n     DIFFERENT POLICY -- per-position, average-cost, ex-fees (V27) -- so "
          "the two\n     disagree row by row BY CONSTRUCTION and agree in total. A "
          "row-level\n     mismatch against the venue is expected here, not a bug in "
          "either.")
    if unscored:
        print(f"  !! settlement unknown, left unscored: {', '.join(unscored)}")
    if provenance:
        print(f"\n  built from PINNED EXPORT: {provenance}"
              "\n  (no venue call — this sheet is reproducible from that file)")
    else:
        print("\n  built from a LIVE venue call — not reproducible; pass "
              "--activities-json\n  with a pinned export if this sheet will be "
              "annotated or re-graded.")
    print(f"\n  xlsx  {xlsx}")
    print(f"  csv   {csv_path}\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
