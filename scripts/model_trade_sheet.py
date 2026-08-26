#!/usr/bin/env python
"""Write the MODEL's side of the annotation workbook — tab 2, beside the
operator's own hand trades in tab 1.

    python scripts/model_trade_sheet.py --tape backups/exports/pulse_decisions_<date>.csv

One row per model ENTRY, joined to its exit, with a plain-language WHY
assembled from the recorded decision context — so the trade can be judged
without reading a tape.

**TAB 1 IS NEVER TOUCHED, AND THE SCRIPT PROVES IT.** Those are the operator's
hand annotations and nothing regenerates them. Before writing, every existing
sheet's cells are snapshotted; after writing, they are re-read and compared. Any
difference restores the backup and raises. A promise not to overwrite is worth
less than a check that it did not.

**DESCRIPTIVE ONLY.** No verdicts, no totals presented as performance. The
registered PULSE measurement lives elsewhere and nothing here may feed it.

Reads the exported tape CSV rather than a database: the export is the artifact
the operator already has, and a sheet-builder that needs a live postgres is one
they cannot run themselves.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import shutil
import sys
from pathlib import Path
from zoneinfo import ZoneInfo

CENTRAL = ZoneInfo("America/Chicago")
MODEL_SHEET = "model trades"

#: (header, width). The last three are the operator's, and stay empty.
COLUMNS: tuple[tuple[str, int], ...] = (
    ("decided (UTC)", 20), ("decided (CT)", 20), ("game", 24),
    ("period", 8), ("clock left", 11), ("score", 10), ("margin", 8),
    ("market", 20), ("side", 6), ("line", 8),
    # Every price below is in the POSITION's own frame, never the venue's YES
    # frame — see _own_frame(). A NO row shows what the NO cost and is worth.
    ("model FV", 10), ("bid", 8), ("ask", 8), ("edge", 8),
    ("contracts", 10), ("stake $", 9), ("live-faithful $", 15),
    ("what set the size", 24),
    ("filled?", 9), ("fill price", 11),
    ("outcome", 20), ("exit price", 11), ("capture per $", 14), ("right?", 9),
    ("WHY — the model's reasoning", 96),
    ("brain", 8),
    ("do I agree", 34), ("what I'd have done", 40), ("pattern name", 24),
)


def _f(value: str | None) -> float | None:
    try:
        return float(value) if value not in (None, "") else None
    except ValueError:
        return None


def _ts(value: str | None) -> dt.datetime | None:
    if not value:
        return None
    text = value.strip().replace(" ", "T", 1)
    if "+" in text and text[-3] == ":":
        pass
    try:
        return dt.datetime.fromisoformat(text)
    except ValueError:
        return None


def _fmt_minutes(value: float | None) -> str:
    return f"{value:.1f}" if value is not None else ""


def _as_int(value) -> int | None:
    return int(float(value)) if value not in (None, "") else None


def _own_frame(row: dict) -> dict:
    """Price, FV and book expressed in the POSITION's own frame.

    The venue quotes everything in the YES frame (V14/V19), so a NO entry
    records limit_price 0.48 meaning "paid 0.52 for the NO". Printed raw beside
    a fair value of 0.31, that reads as the model buying something it valued
    BELOW cost — the exact opposite of the truth, and the same trap the hand
    sheet's market-vs-position columns exist to close. Flip everything for NO.
    """
    yes = (row.get("side") or "").lower() == "yes"
    flip = lambda v: v if (v is None or yes) else 1.0 - v
    bid, ask = _f(row.get("market_bid")), _f(row.get("market_ask"))
    return {
        "price": flip(_f(row.get("limit_price"))),
        "fv": flip(_f(row.get("fair_value"))),
        # The NO side's bid is 1 - the YES ask, so the pair also swaps.
        "bid": bid if yes else (None if ask is None else 1.0 - ask),
        "ask": ask if yes else (None if bid is None else 1.0 - bid),
    }


def _market_label(row: dict) -> str:
    """What was traded, in the words a person uses."""
    mtype = (row.get("sports_market_type") or "").replace(
        "basketball_team_full_game_", "")
    line, side = _f(row.get("line")), (row.get("side") or "").lower()
    if mtype == "total" and line is not None:
        return f"{'OVER' if side == 'yes' else 'UNDER'} {line:g}"
    if mtype == "spread" and line is not None:
        return f"spread {line:+g} ({'yes' if side == 'yes' else 'no'} side)"
    if mtype == "winner":
        return f"moneyline ({'yes' if side == 'yes' else 'no'} side)"
    return mtype or "?"


def _why(row: dict, exit_row: dict | None, capture: float | None) -> str:
    """The sentence. Assembled from recorded facts only — nothing inferred.

    Written so a reader can disagree with it: every clause names the number it
    came from, because "the model liked it" is not something anyone can judge.
    """
    frame = _own_frame(row)
    fv, bid, ask, price = frame["fv"], frame["bid"], frame["ask"], frame["price"]
    edge = _f(row.get("edge_net"))
    mins, margin = _f(row.get("minutes_left")), row.get("margin")
    parts: list[str] = []

    verb = "Entered" if row.get("filled_at") else "Wanted"
    at = f" at {price:.2f}" if price is not None else ""
    parts.append(f"{verb} {_market_label(row)}{at}")

    if fv is not None and bid is not None and ask is not None:
        parts.append(f"model FV {fv:.2f} against a {bid:.2f}/{ask:.2f} book")
    if edge is not None:
        parts.append(f"claimed edge {edge * 100:.1f}c")

    state = []
    if row.get("period"):
        state.append(row["period"])
    if mins is not None:
        est = " est." if (row.get("minutes_left_is_estimate") or "").lower() in ("t", "true") else ""
        state.append(f"{mins:.1f} min left{est}")
    if row.get("score"):
        state.append(f"score {row['score']}")
    if margin not in (None, ""):
        state.append(f"margin {int(float(margin)):+d}")
    if state:
        parts.append("· ".join(state))

    total_so_far, projected = _f(row.get("total_so_far")), _f(row.get("projected_total"))
    if total_so_far is not None and projected is not None:
        parts.append(f"{total_so_far:.0f} pts scored, projecting {projected:.0f}")

    cap = row.get("binding_constraint")
    if cap and cap != "kelly":
        parts.append(f"size set by {cap.replace('_', ' ')}")

    if not row.get("filled_at"):
        parts.append("NEVER FILLED — the order rested and the book left")
    elif exit_row is not None:
        reason = (exit_row.get("reason") or "exit").replace("_", " ")
        px = _own_frame({**exit_row, "side": row.get("side")})["price"]
        held = ""
        a, b = _ts(row.get("filled_at")), _ts(exit_row.get("filled_at"))
        if a and b:
            held = f" after {(b - a).total_seconds() / 60:.0f} min"
        gained = f" for {capture * 100:+.1f}c per $" if capture is not None else ""
        parts.append(f"exited on {reason}{' at ' + format(px, '.2f') if px else ''}{held}{gained}")
    else:
        settled = row.get("settlement")
        if settled in ("0", "1"):
            won = (settled == "1") == ((row.get("side") or "") == "yes")
            parts.append(f"no exit — rode to settlement and {'WON' if won else 'LOST'}")
        else:
            parts.append("no exit — still open at export")
    return "; ".join(parts) + "."


def tape_window(tape: Path) -> tuple[str, str] | None:
    """First and last ``decided_at`` in a tape — the window the sheet covers.

    Stamped into the sheet's own label. A tape is named for the day it was
    EXPORTED, not the period it covers: ``pulse_decisions_20260825.csv`` was
    taken at 18:04 CT and its newest decision is 02:02Z that morning, so a
    reader who takes the filename for the coverage is off by the whole day. The
    window has to be visible on the artifact, not inferred from its name.
    """
    stamps = sorted(r.get("decided_at") or "" for r in csv.DictReader(tape.open()))
    stamps = [x for x in stamps if x]
    return (stamps[0][:19], stamps[-1][:19]) if stamps else None


def build_rows(tape: Path) -> list[list]:
    rows = list(csv.DictReader(tape.open()))
    entries = [r for r in rows if r.get("action") == "enter"]
    # First FILLED exit per entry; an unfilled exit did not end the position.
    exits: dict[str, dict] = {}
    for r in rows:
        if r.get("action") == "exit" and r.get("filled_at") and r.get("entry_id"):
            exits.setdefault(r["entry_id"], r)

    out: list[list] = []
    for e in sorted(entries, key=lambda r: r.get("decided_at") or ""):
        x = exits.get(e.get("id") or "")
        decided = _ts(e.get("decided_at"))
        contracts, stake = _f(e.get("contracts")), _f(e.get("stake_usd"))
        side, settled = (e.get("side") or "").lower(), e.get("settlement")
        frame = _own_frame(e)

        # Capture per dollar staked, in the position's own frame (C11).
        capture = right = None
        if e.get("filled_at") and stake:
            if x is not None:
                px = _f(x.get("limit_price"))
                if px is not None and contracts:
                    proceeds = contracts * (px if side == "yes" else 1 - px)
                    capture = (proceeds - stake) / stake
            elif settled in ("0", "1") and contracts:
                won = (settled == "1") == (side == "yes")
                capture = (contracts * (1.0 if won else 0.0) - stake) / stake
                right = "yes" if won else "no"
        if right is None and capture is not None:
            right = "yes" if capture > 0 else "no"

        if not e.get("filled_at"):
            outcome = "never filled"
        elif x is not None:
            outcome = (x.get("reason") or "exit").replace("_", " ")
        elif settled in ("0", "1"):
            outcome = "rode to settlement"
        else:
            outcome = "open at export"

        out.append([
            decided.strftime("%Y-%m-%d %H:%M:%S") if decided else "",
            decided.astimezone(CENTRAL).strftime("%Y-%m-%d %H:%M:%S") if decided else "",
            e.get("event_slug", ""),
            e.get("period", ""),
            _fmt_minutes(_f(e.get("minutes_left"))),
            e.get("score", ""),
            _as_int(e.get("margin")),
            _market_label(e), side.upper(), _f(e.get("line")),
            frame["fv"], frame["bid"], frame["ask"], _f(e.get("edge_net")),
            round(contracts, 3) if contracts else None,
            round(stake, 2) if stake else None,
            _f(e.get("capped_stake_usd")),
            (e.get("binding_constraint") or "").replace("_", " "),
            "yes" if e.get("filled_at") else "no",
            frame["price"] if e.get("filled_at") else None,
            outcome,
            _own_frame({**x, "side": side})["price"] if x else None,
            round(capture, 4) if capture is not None else None,
            right or "",
            _why(e, x, capture),
            e.get("estimates_version", ""),
            "", "", "",
        ])
    return out


def write_sheet(workbook: Path, rows: list[list], *, label: str) -> dict:
    """Replace the model sheet, and PROVE every other sheet is untouched."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    if not workbook.exists():
        raise SystemExit(f"workbook not found: {workbook}")

    backup = workbook.with_name(
        f"{workbook.stem}.backup-"
        f"{dt.datetime.now(dt.timezone.utc):%Y%m%dT%H%M%SZ}{workbook.suffix}")
    shutil.copy2(workbook, backup)

    wb = load_workbook(workbook)
    protected = [s for s in wb.sheetnames if s != MODEL_SHEET]
    before = {
        name: [[c.value for c in row] for row in wb[name].iter_rows()]
        for name in protected
    }

    if MODEL_SHEET in wb.sheetnames:
        del wb[MODEL_SHEET]                      # idempotent regeneration
    ws = wb.create_sheet(MODEL_SHEET)

    ws.append([f"MODEL TRADES — descriptive only, no verdict. Source: {label}"])
    ws["A1"].font = Font(bold=True, italic=True)
    # A MIXED 'brain' COLUMN WITHIN ONE GAME IS THE DESIGN, NOT A BUG. The
    # label records what priced THAT row: from v4 on, a totals row priced by
    # the pace/efficiency decomposition says v4, a winner row with an
    # availability flag active says v4, and an unflagged winner row prices
    # identically to v3 and therefore says v3. Without this note the operator
    # would reasonably read the mixture as corruption and file it as one.
    ws.append([("'brain' is per ROW, not per game — a game can show v3 and v4 "
                "rows together. The label records what priced that row; an "
                "unflagged winner prices as v3 even under v4.")])
    ws["A2"].font = Font(italic=True)
    ws.append([h for h, _ in COLUMNS])
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for i, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in rows:
        ws.append(row)
    why_col = get_column_letter([h for h, _ in COLUMNS].index("WHY — the model's reasoning") + 1)
    for r in range(4, ws.max_row + 1):
        ws[f"{why_col}{r}"].alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(workbook)

    # The proof. Re-read from disk; a promise is not a check.
    verify = load_workbook(workbook)
    for name in protected:
        if name not in verify.sheetnames:
            shutil.copy2(backup, workbook)
            raise SystemExit(f"REFUSING: sheet {name!r} vanished; backup restored")
        after = [[c.value for c in row] for row in verify[name].iter_rows()]
        if after != before[name]:
            shutil.copy2(backup, workbook)
            raise SystemExit(
                f"REFUSING: sheet {name!r} changed; backup restored from {backup.name}")
    return {"protected": protected, "backup": backup, "rows": len(rows)}


def main() -> int:
    parser = argparse.ArgumentParser(prog="meridian-model-trade-sheet")
    parser.add_argument("--tape", type=Path, required=True,
                        help="exported pulse_decisions CSV")
    parser.add_argument("--workbook", type=Path, default=Path(
        "/Users/yayardia/Documents/Quant/Meridian/backups/exports/"
        "wnba-trades-2026-08-17.xlsx"))
    args = parser.parse_args()

    rows = build_rows(args.tape)
    if not rows:
        print("no model entries in that tape — nothing written", file=sys.stderr)
        return 1
    window = tape_window(args.tape)
    covers = f" · covers {window[0]}Z → {window[1]}Z" if window else " · coverage UNKNOWN"
    label = (f"{args.tape.name} · {len(rows)} model entries{covers}"
             f" · written {dt.datetime.now(dt.timezone.utc):%Y-%m-%d %H:%MZ}")
    info = write_sheet(args.workbook, rows, label=label)

    filled = sum(1 for r in rows if r[18] == "yes")
    print(f"\n  wrote '{MODEL_SHEET}': {info['rows']} model entries ({filled} filled)")
    if window:
        print(f"  tape covers: {window[0]}Z -> {window[1]}Z  "
              f"(the FILENAME is the export date, not the coverage)")
    print(f"  protected sheets verified unchanged: {', '.join(info['protected'])}")
    print(f"  backup: {info['backup'].name}")
    print(f"  workbook: {args.workbook}\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
